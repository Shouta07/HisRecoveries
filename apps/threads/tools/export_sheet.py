"""投稿管理用スプレッドシート(.xlsx)を生成する。

現在の accounts/<id>/ の内容（連投テンプレ・引用リソース・ネタ・設定）を、
Google Sheets にアップロードしてそのまま編集できる多タブ構成で書き出す。

    python tools/export_sheet.py mens-body-lab out.xlsx

openpyxl は遅延インポート（未インストールでも他機能に影響しない）。
"""

import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent


def _load(acc: Path, name: str, default):
    p = acc / name
    if not p.exists():
        return default
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return default


def build(account_id: str, out_path: str) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    acc = ROOT / "accounts" / account_id
    persona = _load(acc, "persona.json", {})
    hyp = _load(acc, "hypotheses.json", {"hypotheses": []})
    tmpl = _load(acc, "thread_templates.json", {})
    csrc = _load(acc, "content_sources.json", {"sources": []})

    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="2B2B29")
    head_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = head_fill
            cell.font = head_font

    def autosize(ws, widths):
        from openpyxl.utils import get_column_letter
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── 使い方 ──
    ws = wb.active
    ws.title = "使い方"
    lines = [
        ["His Recoveries Threads 投稿管理シート"],
        [""],
        ["このシートで編集できるもの:"],
        ["・連投テンプレ … 各ネタの本文（本文列を直接書き換え）"],
        ["・引用リソース … 記事/体験メモ由来の14投稿（本文列を編集）"],
        ["・ネタ一覧 … 状態を active/paused にして 出す/止める を切替"],
        ["・設定 … 投稿時刻・比率・時間帯の出し分け"],
        [""],
        ["ルール(厳守): 売り込まない/煽らない/効果断定しない/絵文字0〜2/ハッシュタグ0〜1"],
        ["ギフトは『応援』であって『指摘』ではない。完全守秘（内容は本人以外に知らされない）を安心材料に。"],
        [""],
        ["自動反映するには: サービスアカウントを作成しこのシートを共有 → GitHub Secrets に"],
        ["GOOGLE_SHEETS_ID と GOOGLE_SHEETS_CREDENTIALS_JSON を設定（詳細はREADME）。"],
    ]
    for r in lines:
        ws.append(r)
    ws["A1"].font = Font(bold=True, size=14)
    autosize(ws, [110])

    # ── 連投テンプレ ──
    ws = wb.create_sheet("連投テンプレ")
    ws.append(["ネタID", "種別", "順番", "本文", "備考"])
    type_of = {h["id"]: h.get("type", "") for h in hyp.get("hypotheses", [])}
    slug_type = {h.get("slug", h["id"]): h.get("type", "") for h in hyp.get("hypotheses", [])}
    for slug, posts in tmpl.items():
        t = slug_type.get(slug, "")
        for i, body in enumerate(posts, 1):
            note = "CTA(リンクはここ)" if "{link}" in body else ""
            ws.append([slug, t, i, body, note])
    style_header(ws, 5)
    autosize(ws, [22, 10, 6, 70, 16])
    for row in ws.iter_rows(min_row=2):
        row[3].alignment = wrap

    # ── 引用リソース ──
    ws = wb.create_sheet("引用リソース")
    ws.append(["リソースID", "分類", "順番", "本文", "リンク(誘導のみ)"])
    order = ["empathy", "insight", "question", "fact", "core"]
    label = {"empathy": "共感", "insight": "気づき", "question": "問題提起",
             "fact": "実績", "core": "誘導(core)"}
    for s in csrc.get("sources", []):
        sid = s.get("id", "")
        url = s.get("url", "")
        mats = s.get("materials", {})
        for cat in order:
            items = mats.get(cat)
            if cat == "core":
                if items:
                    ws.append([sid, label[cat], 1, items, url])
            else:
                for i, body in enumerate(items or [], 1):
                    ws.append([sid, label[cat], i, body, ""])
    style_header(ws, 5)
    autosize(ws, [18, 12, 6, 70, 40])
    for row in ws.iter_rows(min_row=2):
        row[3].alignment = wrap

    # ── ネタ一覧・ON/OFF ──
    ws = wb.create_sheet("ネタ一覧")
    ws.append(["ネタID", "種別", "名前", "状態(active/paused)"])
    for h in hyp.get("hypotheses", []):
        ws.append([h["id"], h.get("type", ""), h.get("name", ""),
                   h.get("status", "active")])
    style_header(ws, 4)
    autosize(ws, [22, 10, 30, 22])

    # ── 設定 ──
    ws = wb.create_sheet("設定")
    ws.append(["項目", "値", "説明"])
    p = persona.get("posting", {})
    slot = p.get("slot_type_map", {})
    rows = [
        ["投稿時刻", " / ".join(p.get("time_slots", [])), "1日の配信時刻(JST)"],
        ["1日の本数", p.get("frequency_per_day", ""), ""],
        ["引用リソース比率", p.get("source_post_ratio", ""), "0=連投のみ 1=引用のみ"],
        ["朝の対象層", ", ".join(slot.get("morning", [])), "朝9時に出すtype"],
        ["夜の対象層", ", ".join(slot.get("night", [])), "夜21時に出すtype"],
    ]
    for r in rows:
        ws.append(r)
    style_header(ws, 3)
    autosize(ws, [20, 30, 40])

    # ── 分析(閲覧数) ──
    ws = wb.create_sheet("分析")
    ws.append(["日時", "ネタ", "閲覧数", "いいね", "返信", "リポスト", "エンゲ率", "リンク先"])
    ws.append(["※ collect(GitHub Actions)が投稿後に自動で書き込みます", "", "", "", "", "", "", ""])
    style_header(ws, 8)
    autosize(ws, [18, 24, 10, 8, 8, 8, 10, 20])

    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    account = sys.argv[1] if len(sys.argv) > 1 else "mens-body-lab"
    out = sys.argv[2] if len(sys.argv) > 2 else "投稿管理.xlsx"
    path = build(account, out)
    print("wrote", path)
